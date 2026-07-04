import openpyxl
import re
import sys
import os
from openpyxl.utils import range_boundaries
from copy import copy

# Base directory for the skill to find assets
SKILL_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_TEMPLATE = os.path.join(SKILL_ROOT, 'assets', 'template_aesp.xlsx')

def get_template_path():
    """Returns the internal template path."""
    return DEFAULT_TEMPLATE

def parse_sql(sql_content):
    entries = []
    # Remover comentários de bloco e linha para evitar falsos positivos
    sql_content = re.sub(r'/\*.*?\*/', '', sql_content, flags=re.DOTALL)
    sql_content = re.sub(r'--.*$', '', sql_content, flags=re.MULTILINE)
    
    # Normalizar espaços
    sql_content = re.sub(r'\s+', ' ', sql_content).strip()
    
    # Dividir por ponto e vírgula, mas ignorar nulos
    raw_statements = [s.strip() for s in sql_content.split(';') if s.strip()]
    seen_patterns = {}
    
    for stmt in raw_statements:
        # 1. INSERT INTO ... SELECT (Deve ser tratado como INSERT na aba EE)
        match_insert_select = re.search(r"INSERT\s+INTO\s+([\w\.]+)\s*\((.*?)\)\s*SELECT\s+(.*?)\s+FROM\s+([\w\.]+)", stmt, re.IGNORECASE)
        if match_insert_select:
            table = match_insert_select.group(1)
            stmt_cols = [c.strip() for c in match_insert_select.group(2).split(',')]
            pattern = f"INSERT_SELECT|{table.lower()}|{match_insert_select.group(2).lower()}"
            if pattern not in seen_patterns:
                seen_patterns[pattern] = True
                entries.append({'type': 'INSERT', 'table': table, 'cols': stmt_cols, 'desc': f"Inclusão em {table}", 'sheet': 'EE'})
            continue

        # 2. INSERT INTO ... VALUES
        match_insert = re.search(r"INSERT\s+INTO\s+([\w\.]+)\s*\((.*?)\)\s*VALUES\s*\(", stmt, re.IGNORECASE)
        if match_insert:
            table = match_insert.group(1)
            stmt_cols = [c.strip() for c in match_insert.group(2).split(',')]
            pattern = f"INSERT|{table.lower()}|{match_insert.group(2).lower()}"
            if pattern not in seen_patterns:
                seen_patterns[pattern] = True
                entries.append({'type': 'INSERT', 'table': table, 'cols': stmt_cols, 'desc': f"Inclusão em {table}", 'sheet': 'EE'})
            continue
            
        # 3. UPDATE
        match_update = re.search(r"UPDATE\s+([\w\.]+)\s+SET\s+(.*?)(?:\s+WHERE\s+(.*))?$", stmt, re.IGNORECASE)
        if match_update:
            table = match_update.group(1)
            set_part, where_part = match_update.group(2), match_update.group(3) if match_update.group(3) else ""
            set_cols = re.findall(r"([\w\.]+)\s*=", set_part)
            where_cols = re.findall(r"([\w\.]+)\s*(?:[=<>!]=?|LIKE|IN|IS|BETWEEN)", where_part, re.IGNORECASE)
            stmt_cols, stmt_seen = [], set()
            for c in set_cols + where_cols:
                if c.lower() not in stmt_seen and not c.isdigit() and c.lower() not in ['null', 'now', 'true', 'false', 'and', 'or']:
                    stmt_cols.append(c); stmt_seen.add(c.lower())
            if not stmt_cols: stmt_cols = ['A_PREENCHER_PELO_USUARIO']
            cols_pattern = ",".join(sorted([c.lower() for c in stmt_cols]))
            pattern = f"UPDATE|{table.lower()}|{cols_pattern}"
            if pattern not in seen_patterns:
                seen_patterns[pattern] = True
                entries.append({'type': 'UPDATE', 'table': table, 'cols': stmt_cols, 'desc': f"Alteração em {table}", 'sheet': 'EE'})
            continue

        # 4. DELETE
        match_delete = re.search(r"DELETE\s+FROM\s+([\w\.]+)(?:\s+WHERE\s+(.*))?$", stmt, re.IGNORECASE)
        if match_delete:
            table = match_delete.group(1)
            where_part = match_delete.group(2) if match_delete.group(2) else ""
            where_cols = re.findall(r"([\w\.]+)\s*(?:[=<>!]=?|LIKE|IN|IS|BETWEEN)", where_part, re.IGNORECASE)
            stmt_cols, stmt_seen = [], set()
            for c in where_cols:
                if c.lower() not in stmt_seen and not c.isdigit() and c.lower() not in ['null', 'now', 'true', 'false', 'and', 'or']:
                    stmt_cols.append(c); stmt_seen.add(c.lower())
            if not stmt_cols: stmt_cols = ['A_PREENCHER_PELO_USUARIO']
            cols_pattern = ",".join(sorted([c.lower() for c in stmt_cols]))
            pattern = f"DELETE|{table.lower()}|{cols_pattern}"
            if pattern not in seen_patterns:
                seen_patterns[pattern] = True
                entries.append({'type': 'DELETE', 'table': table, 'cols': stmt_cols, 'desc': f"Exclusão em {table}", 'sheet': 'EE'})
            continue

        # 5. SELECT Autônomo (Somente se não for parte de um INSERT)
        match_select = re.search(r"^SELECT\s+(.*?)\s+FROM\s+([\w\.]+)(?:\s+WHERE\s+(.*))?$", stmt, re.IGNORECASE)
        if match_select:
            select_part, table, where_part = match_select.group(1), match_select.group(2), match_select.group(3) if match_select.group(3) else ""
            is_se = any(k in select_part.lower() for k in ['count(', 'sum(', 'avg(', 'max(', 'min(', '+', '-', '*', '/'])
            stmt_cols, stmt_seen, is_star = [], set(), select_part.strip() == "*"
            if not is_star:
                for c in re.findall(r"([\w\.]+)", select_part):
                    if c.lower() not in stmt_seen and c.lower() not in ['as', 'distinct', 'all', 'select', 'from', 'where'] and not c.isdigit():
                        stmt_cols.append(c); stmt_seen.add(c.lower())
            for c in re.findall(r"([\w\.]+)\s*(?:[=<>!]=?|LIKE|IN|IS|BETWEEN)", where_part, re.IGNORECASE):
                if c.lower() not in stmt_seen and not c.isdigit() and c.lower() not in ['null', 'now', 'and', 'or', 'true', 'false', 'select', 'from', 'where']:
                    stmt_cols.append(c); stmt_seen.add(c.lower())
            if is_star: stmt_cols.append('A_PREENCHER_PELO_USUARIO (SELECT *)')
            elif not stmt_cols: stmt_cols = ['A_PREENCHER_PELO_USUARIO']
            cols_pattern = ",".join(sorted([c.lower() for c in stmt_cols]))
            pattern = f"SELECT|{table.lower()}|{cols_pattern}"
            if pattern not in seen_patterns:
                seen_patterns[pattern] = True
                entries.append({'type': 'SELECT', 'table': table, 'cols': stmt_cols, 'desc': f"Consulta em {table}", 'sheet': 'SE' if is_se else 'CE'})
            continue
    return entries

def copy_row_attributes(ws, src_row, target_row):
    """Copies format and adapts formulas from src_row to target_row."""
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=src_row, column=col)
        new_cell = ws.cell(row=target_row, column=col)
        
        # Copy style (font, border, fill, etc.)
        if source_cell.has_style:
            new_cell.font = copy(source_cell.font)
            new_cell.border = copy(source_cell.border)
            new_cell.fill = copy(source_cell.fill)
            new_cell.number_format = copy(source_cell.number_format)
            new_cell.protection = copy(source_cell.protection)
            new_cell.alignment = copy(source_cell.alignment)
        
        # Copy and adapt formula
        val = source_cell.value
        if val and isinstance(val, str) and val.startswith('='):
            row_shift = target_row - src_row
            # Regex to find row numbers in formulas and increment them by row_shift
            formula = re.sub(r'([A-Z]+)(\d+)', 
                            lambda m: f"{m.group(1)}{int(m.group(2)) + row_shift}" if int(m.group(2)) == src_row else m.group(0), 
                            val)
            new_cell.value = formula
        elif val and col not in [2, 9, 11, 13]: # Don't copy data values (Column 13 is M)
            new_cell.value = val

def fill_spreadsheet(file_path, entries_data):
    wb = openpyxl.load_workbook(file_path)

    # 0. Especialização AESP: Preencher Sumário
    if 'Sumário' in wb.sheetnames:
        ws_sum = wb['Sumário']
        # Célula padrão para Tipo da Demanda no template consolidado
        ws_sum.cell(row=11, column=7).value = 'Apuração Especial (AESP)'

    sheet_groups = {}
    for entry in entries_data:
        s = entry.get('sheet', 'EE')
        if s not in sheet_groups: sheet_groups[s] = []
        sheet_groups[s].append(entry)

    for sheet_name, entries in sheet_groups.items():
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]

        # 1. Identify start row and count existing empty rows
        # Most templates have pre-formatted rows starting at 19
        start_row = 19
        available_rows = 0
        for r in range(start_row, 100):
            # Check if Row exists and is either empty or just has 'Desenvolvimento' in Col M
            name_val = ws.cell(row=r, column=2).value
            type_label = ws.cell(row=r, column=5).value
            # Summary labels usually start here
            if any(k in str(name_val or "") or k in str(type_label or "") for k in ['Qtd de ', 'Total', 'Complexidade']):
                break
            available_rows += 1

        # 2. Insert only if necessary
        if len(entries) > available_rows:
            num_to_insert = len(entries) - available_rows + 2
            insertion_point = start_row + available_rows
            print(f"Inserting {num_to_insert} additional rows at {insertion_point} in {sheet_name}")
            ws.insert_rows(insertion_point, amount=num_to_insert)
            # Copy template to new rows
            for r in range(insertion_point, insertion_point + num_to_insert):
                copy_row_attributes(ws, start_row, r)

        # 3. Fill entries
        for i, data in enumerate(entries):
            row = start_row + i

            # Fill data
            ws.cell(row=row, column=2).value = data.get('desc')

            # Use provided demand_type, or keep 'Desenvolvimento' if it exists and no type provided
            d_type = data.get('demand_type')
            if d_type:
                ws.cell(row=row, column=13).value = d_type
            elif not ws.cell(row=row, column=13).value:
                ws.cell(row=row, column=13).value = 'Desenvolvimento'

            # Memória TD: REMOVE SPACES after commas
            cols = [c.strip() for c in data.get('cols', [])]
            ws.cell(row=row, column=11).value = ",".join(cols)

            # TD count (Column I)
            if cols and 'A_PREENCHER' in cols[0]:
                ws.cell(row=row, column=9).value = 1
            else:
                ws.cell(row=row, column=9).value = len(cols)

            # Re-apply merge B:C if not already there
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)

    wb.save(file_path)
    print(f"Successfully updated {len(entries_data)} entries.")

if __name__ == "__main__":
    pass
